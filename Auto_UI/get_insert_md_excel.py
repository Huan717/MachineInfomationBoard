import openpyxl as opxl
import datetime as dt

from sql_tool import *

'''         Excel 座標            '''
#    存座標 , 行 , 列
class ex_coordinate:
    def __init__(self,coordinate=None,row=None,col=None):
        self.coordinate=coordinate
        self.row=row
        self.col=col
#    存開始的格子 , 結束格子
class arr_range_size:
    def __init__(self,col_s=None,col_e=None,row_s=None,row_e=None):
        self.col_s=col_s
        self.col_e=col_e
        self.row_s=row_s
        self.row_e=row_e
'''      預定表　ＭＡＳＫ　ＤＡＴＡ　ＶＡＬＵＥ            '''


class mask_value:
    #    存符合條件之值
    def __init__(self,
                 lot        = '',
                 maskid     = '',
                 customer   ='',
                 size       ='',
                 masktype   ='binary',
                 machine    ='',
                 start_time ='',
                 end_time   ='',
                 pornot='',
                 data_start='',
                 data_end=''):
        self.lot = lot
        self.maskid = maskid
        self.customer = customer
        self.size = size
        self.masktype = masktype    #是否為 HTM
        self.machine=machine
        self.start_time = start_time
        self.end_time = end_time
        self.pornot=pornot
        self.data_start=data_start
        self.data_end=data_end


'''-----    輔助工具       -----'''
def Machine_check(str):
    if str=='Machine1' or str== '(P)Machine1' or str=='Machine2' or str=='(P)Machine2':
        return True
    else:
        return False
def input_Mask_info(exdb,start,end):
    #print(end.coordinate)
    data_1={
        'data_start' : start.coordinate,
        'data_end'   : end.coordinate,
        'lot'        : str(exdb['C' + str(start.row+1)].value),
        'maskid'     : exdb['C' + str(start.row)].value,
        'customer'   : (exdb['C' + str(start.row-1)].value.split())[0],
        'size'       : (exdb['C' + str(start.row - 1)].value.split())[1],
        'start_time' : str(exdb.cell(row=10, column=start.column).value)[0:16],
        'end_time'   : str(exdb.cell(row=10, column=end.column+1).value)[0:16],
        'status'      : '' ,
        'masktype'   : 'binary',
        'Machine'    : '',
        'pornot'     : 'NP',
    }
    data=mask_value()
    data.data_start =start.coordinate
    data.data_end   =end.coordinate
    data.lot        =str(exdb['C' + str(start.row+1)].value)
    data.maskid     =exdb['C' + str(start.row)].value
    data.customer   =(exdb['C' + str(start.row-1)].value.split())[0]
    data.size       =(exdb['C' + str(start.row-1)].value.split())[1]
    data.start_time =str(exdb.cell(row=10, column=start.column).value)[0:16]
    data.end_time   =str(exdb.cell(row=10, column=end.column+1).value)[0:16]
    if start.value=='Machine1':
        data.machine='Machine1'
        data_1['Machine']='Machine1'
    elif start.value=='(P)Machine1':
        data.machine = 'Machine1'
        data.pornot='P'
        data_1['Machine'] = 'Machine1'
        data_1['pornot']='P'
    elif start.value=='Machine2':
        data.machine = 'Machine2'
        data_1['Machine'] = 'Machine2'
    else:
        data.machine = 'Machine2'
        data.pornot = 'P'
        data_1['Machine'] = 'Machine2'
        data_1['pornot'] = 'P'
    if  exdb['F' + str(start.row)].value == 'HTM_1st' or \
        exdb['F' + str(start.row)].value == 'HTM_2nd'\
    :
        data.masktype = exdb['F' + str(start.row)].value
        data_1['masktype'] = exdb['F' + str(start.row)].value
    if exdb.cell(row=10, column=end.column).value <= dt.datetime.now():
        data_1['status']='F'
    elif exdb.cell(row=10, column=start.column).value > dt.datetime.now():
        data_1['status'] = 'N'
    else :
        data_1['status'] = 'R'

    #print(data_1)
    return data_1
def lot_to_sql(lot):
    #print("||||11-13",lot[11:13])
    if int(lot[11:13]) < 10:
        return lot[0:11] + '0' + str(int(lot[11:13]) + 1)
    else:
        return lot[0:11] + str(int(lot[11:13] + 1))
'''*******      函式區     *******'''

#       抓取資料
def get_ex_data(now_time,fn):

    try:
        wb_md = opxl.load_workbook(fn)
        ws_md = wb_md["生產計劃&進度"]

        col_str = 26  # -> (→ 列) 開始
        row_str = 33  # -> (↓ 欄) 開始
        range_size = arr_range_size()  # 範圍選取
        # 找到現在時間那"格"
        while (col_str >= 26):
            col_str += 1
            if ((type(ws_md.cell(row=10, column=col_str).value) == dt.datetime) and
                    ws_md.cell(row=10, column=col_str).value.strftime('%Y-%m-%d %H:%M') == now_time.strftime(
                        '%Y-%m-%d %H:00')) \
                    :
                # 列 起點  現在時間 - 72 hr
                col_s_temp = ws_md.cell(row=10, column=col_str - 48)
                range_size.col_s = ex_coordinate(col_s_temp.coordinate.replace("10", ""), col_s_temp.row,
                                                 col_s_temp.column)
                # 列 終點  現在時間 + 72 hr
                col_e_temp = ws_md.cell(row=10, column=col_str + 24)
                range_size.col_e = ex_coordinate(col_e_temp.coordinate.replace("10", ""), col_e_temp.row,
                                                 col_e_temp.column)
                break
            #  找到尚未檢查完的製品
        while (row_str >= 33):
            # 起點 1. 起始值為空 2. valus ≠ 空值 & "完成" 3. 顏色 ≠ 灰色
            if (range_size.row_s == None and
                    ws_md.cell(row=row_str, column=11).value != None and
                    ws_md.cell(row=row_str, column=11).value != "完成" and
                    ws_md.cell(row=row_str, column=11).fill.start_color.rgb != 'FF969696' and
                    ws_md.cell(row=row_str, column=11).fill.start_color.rgb != 'FFFFABE7' and
                    ws_md.cell(row=row_str, column=11).fill.start_color.rgb != 'FF969696') \
                    :
                range_size.row_s = row_str
                # print("range_size.row_str= ", range_size.row_str)
            # 終點 1. valus = 空值 2. 沒有顏色
            elif (ws_md.cell(row=row_str, column=11).value == None and
                  ws_md.cell(row=row_str, column=11).fill.start_color.rgb == '00000000') \
                    :
                # 列 - 4= 最後一格
                range_size.row_e = row_str - 4
                # print("range_size.row_end= ",row_str)
                break
            row_str += 4
        range_bool = True
    except:
        print("find range error")
        range_bool = False

    data_arr = []
    start = None
    # 撈取資料
    if range_bool:
        print(range_size.col_s.coordinate, " | ", range_size.col_e.coordinate, "\n", range_size.row_s, " | ",
              range_size.row_e)
        for i in range(range_size.row_s + 1, range_size.row_e, 4):
            # 換行時 結束資料
            if start != None:
                print(ws_md.cell(row=i - 4, column=range_size.col_e.col).coordinate)
                end = ws_md.cell(row=i - 4, column=range_size.col_e.col)
                data_arr.append(input_Mask_info(ws_md, start, end))
                end = None
                start = None

            for j in range(range_size.col_s.col, range_size.col_e.col):
                if i==638:
                    print("----------------------------------------")
                    print(i,j)
                    print(print(ex_data.fill.start_color.rgb))
                ex_data = ws_md.cell(row=i, column=j)
                #print(ex_data.fill.start_color.rgb)
                if (Machine_check(ex_data.value) and
                        (ex_data.fill.start_color.rgb == 'FF00FF00' or
                         ex_data.fill.start_color.rgb == 'FF00FFFF')) \
                        :  # 綠色(P):FF00FF00  藍色(未P):FF00FFFF
                    # print("aaa")
                    if start == None:
                        start = ex_data
                    elif start != None:
                        end = ws_md.cell(row=i, column=j - 1)
                        data_arr.append(input_Mask_info(ws_md, start, end))
                        # print(start.value, " | ", start.coordinate, end.coordinate)
                        end = None
                        start = ex_data
                elif ((ex_data.fill.start_color.rgb == '00000000' or \
                        ex_data.fill.start_color.rgb == 'FFFFFFFF')and\
                      start != None) \
                        :
                    # print("bbb")
                    end = ws_md.cell(row=i, column=j - 1)
                    print(start,end)
                    data_arr.append(input_Mask_info(ws_md, start, end))
                    end = None
                    start = None

        wb_md.close()
        return data_arr
    else:
        return None

def ex_data_upload(data):
    print("excel data upload  :  Start")
    try:
        md_sql().del_nrf()  # 刪除狀態為 N 的資料
        print("del_ALL")
        for db in data:
            #print(db)
            if db['lot'] != 'None':
                md_sql().lot_insert(db)
                # md_sql().lot_insert(db)
        print("|||||||||||||||||||||||||||||||||||||||||||")
        for db in data:
            temp_count = 0
            sqldb = md_sql().select("SELECT TOP 1 * FROM [dbo].[web_MDKB_No_of_inspect] where lot like '" + db['lot'] + "%' ORDER BY lot desc;")

            #   sql 無 資料 - 直接存入
            if sqldb == []:
                if db['masktype'] == 'HTM-2nd':
                    db['lot'] = db['lot'] + '_201'
                else:
                    db['lot'] = db['lot'] + '_101'
                #print('sqldb==[]',db['lot'])
                md_sql().ex_insert(db)
            else:
                db['lot']=lot_to_sql(sqldb[0][0])
                #print(db['lot'])
                md_sql().ex_insert(db)
            #print("*********************************************")

        print("MD DATA insert finsh")
        print("*********************************************")
    except:
        print("ERROR : MD EXCEL Upload Fail")




